import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os

file_path = os.path.join(".", "spotify-2023.xlsx")

if os.path.exists(file_path): 
    df = pd.read_excel(file_path)
    print(df)
else:
    print(f"The file '{file_path}' does not exist.")

#What is the total number of records in the dataset? #(Finding total row count)
row_count = len(df.axes[0])
#print(f'The DataFrame has {row_count} rows.')

#What is the total number of columns in the dataset? #(Finding total column count)
col_count = len(df.axes[1])
#print(f'The DataFrame has {col_count} columns.')

#What is the range of years in which the tracks were released? #Finding Range of the particular coln
roy_track_min = df['released_year'].min()
roy_track_max = df['released_year'].max()
#print("The range of years in which the tracks were released were", roy_track_min, "-", roy_track_max)

# What is the average number of streams for all tracks?
df['streams'] = pd.to_numeric(df['streams'], errors='coerce')
streams_avg = df["streams"].mean()
# print("The average number of streams for all tracks is:", streams_avg)

# Filtering and Selection:

# How many tracks were released in [specific year] (2022)?
specific_year = 2022
filtered_result = df[df['released_year'] == specific_year] 
#print('\nResult dataframe :\n', filtered_result)

# Show the details of the track with the highest [variable] (danceability_% ).
track_variable  = "danceability_%"
track_variable_max = df[track_variable].max()
#print (track_variable_max)
filtered_highest_track_dance = df[df[track_variable] == track_variable_max] # setting the value to the max
highest_track_name = filtered_highest_track_dance['track_name']
# print("The track(s) with the highest", track_variable, "is/are:")
# print(highest_track_name)

# List the tracks with [specific condition] for [variable]. (More than 2 singer and display the artist name and song)
singer_cnt = 2
singer_tracks = df[df["artist_count"]== singer_cnt]
variable_coln = ["track_name","artist(s)_name"]
tracks_with_2_singer = singer_tracks[variable_coln]
#print("Tracks with more than 2 singers:", tracks_with_2_singer)

# Aggregation:

# What is the total number of streams for all tracks?
streams_sum = df["streams"].sum()
#print ("Total number of streams: ", streams_sum)

# Calculate the average [variable] for tracks in [specific year] (streams,2010). 
specific_year_1 = 2010
filtered_result_1 = df[df['released_year'] == specific_year_1] # set as 2010
tracks_cnt_avg = filtered_result_1["streams"].mean()
# print("The average streams for tracks in 2010 is:", tracks_cnt_avg)



# Find the track with the most streams in [specific playlist type] (Spotify, Apple Music, Deezer). TBC

# highest_stream_cnt_spotify_playlist  = df["in_spotify_playlists"].max() #highest cnt in_spotify_playlist
# highest_streams_spotify_playlist = df[df["streams"] == highest_stream_cnt_spotify_playlist] # streams that has the highest cnt 
# print ("Track with the highest stream in spotify playist is", highest_streams_spotify_playlist)

# # Filter the DataFrame for tracks in the Spotify playlist
# spotify_tracks = df[df['in_spotify_playlists'] == 1]

# # Find the maximum streams within the Spotify playlist
# highest_stream_cnt_spotify_playlist = spotify_tracks['streams'].max()

# # Filter the Spotify subset to find the track with the highest streams
# highest_streams_spotify_playlist = spotify_tracks[spotify_tracks['streams'] == highest_stream_cnt_spotify_playlist]

# # Print the track with the highest streams in the Spotify playlist
# print("Track with the highest stream in the Spotify playlist is:")
# print(highest_streams_spotify_playlist)






# Visualization:

# Create a histogram for the distribution of [variable].
# Plot a histogram for the 'streams' variable
df['streams'].plot.hist(bins=20, edgecolor='k')
plt.xlabel('Streams')
plt.ylabel('Frequency')
plt.title('Distribution of Streams')
# plt.show()

# Plot a line chart to show the trend of [variable] over the years.
# Plot a line chart for the trend of 'streams' over the years
df.groupby('released_year')['streams'].sum().plot(marker='o')
plt.xlabel('Year')
plt.ylabel('Total Streams')
plt.title('Trend of Streams Over the Years')
# plt.show()

# Compare the [variable] between Spotify and Apple Music with a scatter plot.
# Create a scatter plot to compare 'streams' between Spotify and Apple Music
# plt.scatter(df[df['in_spotify_playlists'] == 1]['streams'], df[df['in_apple_playlists'] == 1]['streams'], alpha=0.5)
# plt.xlabel('Spotify Streams')
# plt.ylabel('Apple Music Streams')
# plt.title('Comparison of Streams Between Spotify and Apple Music')
# plt.show() TBC




# Merging and Combining Data:
# Merge the dataset with an additional dataset (if available) and calculate the total streams.
# Combine data from multiple sources (e.g., Spotify, Apple Music, Deezer) into a single DataFrame.


# Data Export and Saving:

# Export track_name, artist_name, streams & released_year to a new Excel files in a new tab.

selected_columns = ['track_name', 'artist(s)_name', 'streams', 'released_year']
filtered_df = df[selected_columns]

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Create a new sheet in the workbook
new_sheet = workbook.active
new_sheet.title = "Selected Columns"

# Write the header row to the new sheet
header_row = selected_columns
new_sheet.append(header_row)

# Write the data from the selected columns to the new sheet
for row in filtered_df.itertuples(index=False):
    new_sheet.append(list(row))

# Save the modified workbook with the new sheet
# workbook.save('output.xlsx')



# Export artist(s)_name in a tab called artist_name

# Create a new DataFrame by filtering the original data
filtered_df1 = df['artist(s)_name']  # Replace 'Your_Artist_Name' with the specific artist name you want to filter

workbook = openpyxl.load_workbook('output.xlsx')

# Create a new DataFrame by extracting the 'artist(s)_name' column
filtered_df1 = df['artist(s)_name']

# Create a new sheet in the workbook
new_sheet1 = workbook.create_sheet("artist_name")

# Add a header row to the new sheet
header_row = ['artist(s)_name']  # Header name for the column
new_sheet1.append(header_row)

# Write the data to the new sheet
for row_num, artist_name in enumerate(filtered_df1, 1):  # Start from row 1
    new_sheet1.cell(row=row_num, column=1, value=artist_name)

# Save the modified workbook with the filtered data in the existing sheet
workbook.save('output.xlsx')


# Save the modified dataset to a new Excel file.

# Export the data to a CSV file for further analysis.

# TBC

# Here are some additional tasks you can consider with pandas and Openpyxl:

# Data Transformation: You can perform data transformations like data cleaning, normalization, and feature engineering, depending on your specific dataset and analysis goals.

# Data Aggregation: If your dataset contains multiple entries for the same track, you can perform aggregation operations, such as finding the total streams for each track or computing the average values for different variables.

# Filtering and Sorting: You can filter data based on specific conditions and sort it by various columns to gain insights.

# Concatenating DataFrames: If you have multiple datasets with similar structures, you can concatenate them into a single DataFrame for analysis.

# Data Visualization: You can create various types of charts and plots to explore data relationships and trends further.

# Data Export: Besides Excel and CSV, you can export data to other formats like JSON, SQLite, or Parquet for different analysis tools.

# Advanced Analysis: If you have specific research questions or analytical goals, you can delve into more advanced statistical analysis or machine learning using libraries like scikit-learn.